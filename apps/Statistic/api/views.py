# NOTE:BookShow 
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment,Border,Side,PatternFill
from openpyxl.utils import get_column_letter
from datetime import date, timedelta,datetime
import calendar

#Django
from django.http import HttpResponse

#DRF
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

#Apps locals
from apps.Utilidades.permisos import Get_Model_Name
from apps.Access.models import UserSession
from apps.Requests.models import Solicitud,Plan
from apps.Access.api.serializers import SerializersUserSession
from apps.Requests.api.serializers import SolicitudSerializers,PlanSerializers

#PErmisos 
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from apps.Utilidades.permisos import RolePermission

class GetStatisticSolicitud(APIView):
    
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, RolePermission]
    allowed_roles = ["AD"]
    
    serializer_class = SolicitudSerializers

    def get(self, request, *args, **kwargs):
        try:

            year = self.kwargs.get('year', None)
            list_moth= list(calendar.month_name)[1:]

            data_count={}
            #Recoremos la lista de los meses con sus indixes 
            for i,month in enumerate(list_moth, start=1):

                #hacemos el Filtro con el año indicado con todos los meses 
                data = Solicitud.objects.filter(fecha__year=year, fecha__month=i, is_active=True)
                information = self.serializer_class(data, many=True).data

                #relacionamos todo con el mes correspodiente
                data_count[month]={
                    'total_solicitudes': len(information),
                    'solicitudes_activo': len([sol for sol in information if sol['estado'] == 'AC']),
                    'solicitudes_cancelado': len([sol for sol in information if sol['estado'] == 'CAL']),
                    'solicitudes_progreso': len([sol for sol in information if sol['estado'] == 'PRO']),
                    'solicitudes_finalizado': len([sol for sol in information if sol['estado'] == 'FIN']),
                    }
                
            return Response({"data": data_count}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
#NOTE: La Vista esta terminada pendiente que frontend haga pruebas necesarias
class GETPlanes(APIView):
    
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, RolePermission]
    allowed_roles = ["AD"]
    
    serializer_class = PlanSerializers

    def get(self,request):
        try:
            datas = Solicitud.objects.filter(is_active = True)
            column_name = self.serializer_class( Plan.objects.filter(is_active=True), many = True)
            
            data_acount= {}
            for data in column_name.data:
                data_acount[data["nombre_plan"]] = len( [sol for sol in datas if sol.id_plan.pk== data["id"]])
            
            return Response({'data':data_acount}, status=status.HTTP_200_OK)
        
        except Exception as E:
            return Response({"error": str(E)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)    

#NOTE:La vista esta terminada pendiente que frontend haga pruebas necesarias
class ReportLogins(APIView):

    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, RolePermission]
    allowed_roles = ["AD"]
    serializer_class = SerializersUserSession

    def get(self, request):
        try:
            logins = UserSession.objects.all().order_by('-login_count')[:3]
            data_count = {}

            for data in logins:
                usurname = data.id_usuario.usuario  
                data_count[usurname] = data.login_count

            return Response(data_count, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class GetTiempoSolucion(APIView):
         
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, RolePermission]
    allowed_roles = ["AD"]


    model = Solicitud
    serializer_class = SolicitudSerializers

    def get(self, request):
        # Filtrar solicitudes finalizadas del mes actual
        data_create = Solicitud.objects.filter(
            fecha__year=datetime.now().year,
            fecha__month=datetime.now().month,
            estado="FIN"
        )

        total_tiempo = timedelta(0)
        count = 0

        for data in data_create:
            if data.fecha_fin:  
                total_tiempo += (data.fecha_fin - data.fecha)
                count += 1

        if count > 0:  # evitar división por cero
            promedio = total_tiempo / count
            dias = promedio.days
            horas = promedio.seconds // 3600
            promedio_str = f"{dias} días, {horas} horas"
        else:
            promedio_str = "Sin datos"

        return Response({
            "promedio_duracion": promedio_str
        }, status=status.HTTP_200_OK)


class ReportesExcel(APIView):
     
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, RolePermission]
    allowed_roles = ["AD"]

    #Ejempli de URL Que e Debe utilizar 
    #http://localhost:8000/statistic/reporte/?model=convenio&year_start=2024&month_start=6&year_end=2025&month_end=7&state=IN

    def get(self, request, *args, **kwargs):
        # 1. Obtener y validar los datos
        try:
            tabla = str(request.GET.get("model", None))
            year_start = int(request.GET.get("year_start", 2000)) 
            month_start = int(request.GET.get("month_start", 1)) 
            year_end = int(request.GET.get("year_end", int(datetime.now().year))) 
            month_end = int(request.GET.get("month_end", int(datetime.now().month))) 
            state = request.GET.get("state", "Todos")  # valor por defecto
            boolean = request.GET.get("boolean", 0)  # valor por defecto


            print(state)
        except (ValueError, TypeError):
            return HttpResponse("Parámetros inválidos", status=400)
        model = Get_Model_Name(tabla)
        # 2. Construir queryset con filtros
        queryset = model.objects.all()

        if boolean:
            return Response("Datos Verda", status=status.HTTP_200_OK)
        # Filtro por fechas
        if all([year_start, month_start, year_end, month_end]):
            try:
                date_start = date(year_start, month_start, 1)
                # calcular último día del mes final
                if month_end == 12:
                    date_end = date(year_end, month_end, 31)
                else:
                    date_end = date(year_end, month_end + 1, 1) - timedelta(days=1)

                # Detectar dinámicamente un campo de fecha
                fecha_fields = ["fecha", "fecha_creacion", "created_at"]
                campos_modelo = [f.name for f in model._meta.fields]

                campo_fecha = next((f for f in fecha_fields if f in campos_modelo), None)

                if campo_fecha:
                    queryset = queryset.filter(**{f"{campo_fecha}__range": (date_start, date_end)})


            except (ValueError, TypeError):
                return HttpResponse("Rango de fechas inválido", status=400)

        # Filtro por estado
        if state != "Todos":
            queryset = queryset.filter(estado=state)

        # 3. Generar el Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Empleados"
        
        # creacion  de estilos
        borde_fino = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Titulo principal
        campos = [field.name for field in model._meta.fields if field.name not in ['id','password','is_staff','is_superuser']]
        num_cols = len(campos)
        col_fin = get_column_letter(num_cols)
        
        ws.merge_cells(f"A1:{col_fin}1")
        celda_titulo = ws["A1"]
        celda_titulo.value = f"REPORTE DE {tabla.upper()}"
        celda_titulo.font = Font(bold=True, size=14)
        celda_titulo.alignment = Alignment(horizontal="center", vertical="center")

        # Encabezados de columnas
        for col_num, campo in enumerate(campos, 1):
            print(campo)
            if campo == "is_active":
                valor = "Eliminado"
            else:
                valor = campo
            celda=ws.cell(row=2, column=col_num, value=valor.replace('_', ' ').title())
            
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal='center')
            celda.border = borde_fino
            celda.fill= fill

        # Datos
        for row_num, empleado in enumerate(queryset, 3):
            for col_num, campo in enumerate(campos, 1):
                valor = getattr(empleado, campo, None)
                if isinstance(valor, bool):
                    valor = "NO" if valor else "si"
                elif isinstance(valor, date):
                    valor = valor.strftime('%d/%m/%Y')
                elif valor ==None:
                    valor = "N/A"
                celda = ws.cell(row=row_num, column=col_num, value=str(valor))
                celda.border = borde_fino

        # Ajustar ancho de columnas
        for col_num, campo in enumerate(campos, 1):
            max_length = len(campo)
            for row in ws.iter_rows(min_row=3, min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_{tabla}.xlsx"'

        wb.save(response)
        return response